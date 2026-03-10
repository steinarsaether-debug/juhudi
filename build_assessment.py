import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule

# ─────────────────────────────────────────────────────────────
# COLOUR CONSTANTS
# ─────────────────────────────────────────────────────────────
AN_DARK_BLUE = "1B3A6B"
AN_LIGHT     = "D9E1F2"
SEC_HEADER   = "2F5496"
ROW_ALT      = "F2F2F2"
ROW_WHITE    = "FFFFFF"
SOLID_BG     = "E2EFDA"
LOW_BG       = "FFE7E7"
RED_SCORE    = "C00000"
AMBER_SCORE  = "FF8C00"
GREEN_BG     = "92D050"
GREEN_TEXT   = "375623"
BLACK        = "000000"
WHITE        = "FFFFFF"
CONTEXT_BG   = "EAF0FB"

# ─────────────────────────────────────────────────────────────
# STYLE HELPERS
# ─────────────────────────────────────────────────────────────
def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def afont(bold=False, color=BLACK, size=11, italic=False):
    return Font(bold=bold, color=color, size=size, italic=italic, name="Calibri")

def center_wrap():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def left_top():
    return Alignment(horizontal="left", vertical="top", wrap_text=True)

def left_mid():
    return Alignment(horizontal="left", vertical="center", wrap_text=True)

def right_mid():
    return Alignment(horizontal="right", vertical="center", wrap_text=False)

def thin_border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def med_border():
    s = Side(style="medium", color="888888")
    return Border(left=s, right=s, top=s, bottom=s)

def set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def rh(ws, row, h):
    ws.row_dimensions[row].height = h

# ─────────────────────────────────────────────────────────────
# WRITE TO FIRST CELL OF A MERGED RANGE
# (merge first, then write to top-left cell)
# ─────────────────────────────────────────────────────────────
def merged_cell(ws, r1, c1, r2, c2):
    ws.merge_cells(
        start_row=r1, start_column=c1,
        end_row=r2, end_column=c2
    )
    return ws.cell(row=r1, column=c1)

# ─────────────────────────────────────────────────────────────
# CONDITIONAL FORMATTING HELPER
# ─────────────────────────────────────────────────────────────
def add_score_cf(ws, cell_ref):
    """Apply red/amber/green CF to a single cell address string."""
    red_f  = PatternFill(start_color=RED_SCORE,   end_color=RED_SCORE,   fill_type="solid")
    amb_f  = PatternFill(start_color=AMBER_SCORE,  end_color=AMBER_SCORE, fill_type="solid")
    grn_f  = PatternFill(start_color=GREEN_BG,     end_color=GREEN_BG,    fill_type="solid")
    red_fn = Font(bold=True, color=WHITE,      name="Calibri", size=11)
    amb_fn = Font(bold=True, color=WHITE,      name="Calibri", size=11)
    grn_fn = Font(bold=True, color=GREEN_TEXT, name="Calibri", size=11)
    ws.conditional_formatting.add(cell_ref,
        CellIsRule(operator="lessThan",           formula=["0.4"],          fill=red_f, font=red_fn))
    ws.conditional_formatting.add(cell_ref,
        CellIsRule(operator="between",            formula=["0.4","0.6999"], fill=amb_f, font=amb_fn))
    ws.conditional_formatting.add(cell_ref,
        CellIsRule(operator="greaterThanOrEqual", formula=["0.7"],          fill=grn_f, font=grn_fn))

def add_row_highlight_cf(ws, row):
    """Highlight full row when answer is 0 or 1."""
    rng = f"A{row}:G{row}"
    ws.conditional_formatting.add(rng,
        CellIsRule(operator="equal", formula=["1"],
                   fill=PatternFill(start_color=SOLID_BG, end_color=SOLID_BG, fill_type="solid")))
    ws.conditional_formatting.add(rng,
        CellIsRule(operator="equal", formula=["0"],
                   fill=PatternFill(start_color=LOW_BG,   end_color=LOW_BG,   fill_type="solid")))

# ─────────────────────────────────────────────────────────────
# SHEET COMPONENT HELPERS
# ─────────────────────────────────────────────────────────────
def write_title_row(ws, row, text, n_cols=7, bg=AN_DARK_BLUE,
                    fg=WHITE, size=16):
    c = merged_cell(ws, row, 1, row, n_cols)
    c.value = text
    c.fill = fill(bg)
    c.font = afont(bold=True, color=fg, size=size)
    c.alignment = center_wrap()
    rh(ws, row, 34)

def write_col_headers(ws, row, headers):
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=ci)
        c.value = h
        c.fill = fill(AN_LIGHT)
        c.font = afont(bold=True, color=AN_DARK_BLUE, size=11)
        c.alignment = center_wrap()
        c.border = thin_border()
    rh(ws, row, 22)

def write_section_header(ws, row, text, context=False):
    c = merged_cell(ws, row, 1, row, 7)
    c.value = text
    c.fill = fill(CONTEXT_BG if context else SEC_HEADER)
    c.font = afont(bold=True,
                   color=(AN_DARK_BLUE if context else WHITE),
                   size=12)
    c.alignment = left_mid()
    c.border = thin_border()
    rh(ws, row, 22)

def write_question_row(ws, row, num, who, question, guidance,
                       score_label, alt=False, context=False):
    bg = CONTEXT_BG if context else (ROW_ALT if alt else ROW_WHITE)
    data = [num, who, question, guidance, None, score_label, ""]
    for ci, val in enumerate(data, 1):
        c = ws.cell(row=row, column=ci)
        c.value = val
        c.fill = fill(bg)
        c.font = afont(size=10)
        c.border = thin_border()
        if ci == 1:
            c.alignment = center_wrap()
        elif ci == 5:
            c.alignment = center_wrap()
        else:
            c.alignment = left_top()
    rh(ws, row, 60)

def write_section_score_row(ws, row, label, avg_formula, context=False):
    c = merged_cell(ws, row, 1, row, 4)
    c.value = label
    c.fill = fill(AN_LIGHT)
    c.font = afont(bold=True, color=AN_DARK_BLUE, size=11)
    c.alignment = right_mid()
    c.border = thin_border()

    sc = ws.cell(row=row, column=5)
    sc.value = avg_formula
    sc.number_format = "0.00"
    sc.font = afont(bold=True, size=11)
    sc.alignment = center_wrap()
    sc.border = med_border()

    for col in (6, 7):
        ws.cell(row=row, column=col).fill = fill(AN_LIGHT)
        ws.cell(row=row, column=col).border = thin_border()
    rh(ws, row, 22)
    return f"E{row}"

def add_answer_dropdown(ws, answer_rows):
    sq = " ".join(f"E{r}" for r in answer_rows)
    dv = DataValidation(type="list", formula1='"0,0.5,1"',
                        allow_blank=True, showDropDown=False)
    dv.sqref = sq
    ws.add_data_validation(dv)

# ─────────────────────────────────────────────────────────────
# SCORE SUMMARY BLOCK
# ─────────────────────────────────────────────────────────────
def write_score_summary(ws, start_row, sections, section_score_refs):
    """
    sections          : list of dicts with 'title' and optional 'context'
    section_score_refs: list of cell references (e.g. 'E42') or None for context
    Returns the row after the block.
    """
    row = start_row
    rh(ws, row, 12); row += 1

    # ── Summary title ──
    c = merged_cell(ws, row, 1, row, 7)
    c.value = "SCORE SUMMARY"
    c.fill = fill(AN_DARK_BLUE)
    c.font = afont(bold=True, color=WHITE, size=13)
    c.alignment = center_wrap()
    rh(ws, row, 26); row += 1

    # ── Column header row ──
    hdr_vals = {1: "Section", 4: "Score", 5: "Rating / Interpretation"}
    for col in range(1, 8):
        c = ws.cell(row=row, column=col)
        c.fill = fill(AN_LIGHT)
        c.font = afont(bold=True, color=AN_DARK_BLUE, size=11)
        c.alignment = center_wrap()
        c.border = thin_border()
        if col in hdr_vals:
            c.value = hdr_vals[col]
    # Merge Section cols 1-3 and Rating cols 5-7
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    ws.merge_cells(start_row=row, start_column=5, end_row=row, end_column=7)
    rh(ws, row, 20); row += 1

    # ── One row per section ──
    scored_refs = []
    for i, (sec, ref) in enumerate(zip(sections, section_score_refs)):
        is_ctx = sec.get("context", False)
        bg = CONTEXT_BG if is_ctx else (ROW_ALT if i % 2 else ROW_WHITE)

        # Cols 1-3: section name
        c = merged_cell(ws, row, 1, row, 3)
        c.value = sec["title"]
        c.fill = fill(bg)
        c.font = afont(size=11, italic=is_ctx,
                       color=AN_DARK_BLUE if is_ctx else BLACK)
        c.alignment = left_mid()
        c.border = thin_border()

        # Col 4: score (formula link)
        sc = ws.cell(row=row, column=4)
        if ref and not is_ctx:
            sc.value = f"={ref}"
            sc.number_format = "0.00"
            sc.font = afont(bold=True, size=11)
            sc.alignment = center_wrap()
            sc.border = med_border()
            add_score_cf(ws, f"D{row}")
            scored_refs.append(f"D{row}")
        else:
            sc.value = "—"
            sc.fill = fill(bg)
            sc.font = afont(size=10, italic=True)
            sc.alignment = center_wrap()
            sc.border = thin_border()

        # Cols 5-7: rating
        rc = merged_cell(ws, row, 5, row, 7)
        if is_ctx:
            rc.value = "Context only — not included in overall score"
            rc.font = afont(size=10, italic=True, color=AN_DARK_BLUE)
        elif ref:
            rc.value = (f'=IF({ref}<0.4,"High Risk — significant investment needed",'
                        f'IF({ref}<0.7,"Moderate — gaps require attention",'
                        f'"Strong — above average for peer group"))')
            rc.font = afont(size=10)
        rc.fill = fill(bg)
        rc.alignment = left_mid()
        rc.border = thin_border()
        rh(ws, row, 22); row += 1

    # ── Overall score row ──
    rh(ws, row, 6); row += 1

    non_ctx_refs = [r for r, s in zip(section_score_refs, sections)
                    if r and not s.get("context", False)]
    avg_f = f"=IFERROR(AVERAGE({','.join(non_ctx_refs)}),\"\")"

    c = merged_cell(ws, row, 1, row, 3)
    c.value = "OVERALL DIGITAL MATURITY SCORE"
    c.fill = fill(AN_DARK_BLUE)
    c.font = afont(bold=True, color=WHITE, size=13)
    c.alignment = right_mid()
    c.border = med_border()

    ov = ws.cell(row=row, column=4)
    ov.value = avg_f
    ov.number_format = "0.00"
    ov.font = afont(bold=True, size=14)
    ov.alignment = center_wrap()
    ov.border = med_border()
    add_score_cf(ws, f"D{row}")
    overall_ref = f"D{row}"

    ic = merged_cell(ws, row, 5, row, 7)
    ic.value = (f'=IF({overall_ref}<0.4,'
                f'"High digital risk — significant investment in technology may be needed",'
                f'IF({overall_ref}<0.7,'
                f'"Moderate — digital foundations exist but gaps require attention",'
                f'"Strong — solid digital maturity, above average for peer group"))')
    ic.font = afont(bold=True, size=11)
    ic.alignment = left_top()
    ic.border = med_border()
    rh(ws, row, 34); row += 1

    return row


# ─────────────────────────────────────────────────────────────
# PRELIM QUESTION DATA
# ─────────────────────────────────────────────────────────────
PRELIM_SECTIONS = [
    {
        "title": "SECTION 1: Digital Strategy & Leadership",
        "questions": [
            ("Q1","CEO / COO",
             "Does the institution have a written digital transformation strategy, approved by the board, with a clear roadmap of at least 3 years?",
             "Look for: board approval, link to business goals, concrete milestones. Ask for the document.",
             "Low: No strategy or just a vague plan\nPartial: Strategy exists but not board-approved or lacks roadmap\nSolid: Board-approved strategy with 3-year roadmap tied to business goals"),
            ("Q2","CEO / COO",
             "What share of the annual operating budget is allocated to technology and digital initiatives? Has this increased in the past 2 years?",
             "<2% = Low; 2-5% = Partial; >5% or growing significantly = Solid",
             "Low: <2% or unknown\nPartial: 2–5%, stable\nSolid: >5% or growing year-on-year"),
            ("Q3","CEO / COO",
             "Who is your most senior technology leader, and who do they report to?",
             "CTO/CIO reporting to CEO = Solid. IT Manager buried in operations = Low.",
             "Low: IT Manager reporting 2+ levels below CEO\nPartial: IT Director, one level below CEO\nSolid: CTO or CIO reporting directly to CEO"),
            ("Q4","CEO / COO",
             "Are there digital KPIs tracked and reported to management/board regularly (e.g. % digital transactions, system uptime, digital onboarding rate)?",
             "Ask for examples of KPIs. A dashboard would be positive.",
             "Low: No digital KPIs\nPartial: KPIs exist but tracked informally\nSolid: Dashboard with digital KPIs reported to board/management"),
        ]
    },
    {
        "title": "SECTION 2: Core Banking System",
        "questions": [
            ("Q5","CTO / CEO",
             "What is the name of your core banking system (CBS)? Is it hosted in the cloud, at a third-party data center, or on your own servers?",
             "Cloud-native (Mambu, Oradian, Musoni, etc.) = Solid. Own server room with aging hardware = Low.",
             "Low: Unknown/legacy in-house system or own server room, hardware >5 years old\nPartial: CBS at third-party hosted data center, or standard platform with on-premise deployment\nSolid: Cloud-native SaaS CBS (e.g. Mambu, Oradian, Musoni, Craft Silicon cloud)"),
            ("Q6","CTO / CEO",
             "How long have you been on your current CBS, and are you planning any upgrade or migration?",
             "Older than 10 years without upgrade = Low. Recent upgrade or cloud migration planned = Solid.",
             "Low: CBS implemented >10 years ago with no upgrade planned\nPartial: CBS 5–10 years old, some upgrade discussion\nSolid: CBS <5 years old or recently migrated/upgraded"),
            ("Q7","CTO / CEO",
             "Does your core banking system process transactions and generate reports in real time, or is there a daily batch processing cycle?",
             "Real-time = Solid. End-of-day batch = Low.",
             "Low: All reports require manual extraction or end-of-day batch run\nPartial: Core transactions real-time but reporting still requires manual steps\nSolid: Full real-time processing and on-demand reporting"),
            ("Q8","CTO / CEO",
             "Does your CBS expose APIs that allow it to connect with other systems (e.g. mobile money, credit bureaus, fintech partners)?",
             "Ask: can they connect to M-Pesa / MTN MoMo via API? Direct integration vs. file transfers matter.",
             "Low: No APIs or all integrations done via manual file transfers\nPartial: Some APIs exist but limited or poorly documented\nSolid: Documented REST API layer enabling real-time integrations with third parties"),
        ]
    },
    {
        "title": "SECTION 3: Digital Channels & Mobile Money",
        "questions": [
            ("Q9","CEO / CCO",
             "What percentage of loan repayments are received through digital channels (mobile money, bank transfer, app)? What percentage are still cash collected?",
             "<10% digital = Low. 10-50% = Partial. >50% = Solid. An institution that cannot measure this = Low.",
             "Low: <10% digital repayments, or unable to measure\nPartial: 10–50% of repayments via digital channels\nSolid: >50% of repayments via digital channels"),
            ("Q10","CEO / CCO",
             "Is your institution integrated with mobile money providers (e.g. M-Pesa, MTN MoMo, Airtel Money, Wave)? Can customers receive disbursements and make repayments via mobile money?",
             "Integration = Solid. No integration in a market with high mobile money penetration = Low.",
             "Low: No mobile money integration in a market where it is widely available\nPartial: Integration with one provider, limited to one transaction type\nSolid: Integrated with main provider(s), both disbursement and repayment supported"),
            ("Q11","CEO / CCO",
             "Can a customer apply for a loan or open an account digitally, without visiting a branch? If so, what does the process look like?",
             "Fully digital onboarding = Solid. Branch visit required for everything = Low.",
             "Low: All customers must visit a branch for all steps\nPartial: Some steps digital (e.g. application form online) but branch visit still required for approval or ID verification\nSolid: End-to-end digital onboarding available, including digital ID verification (eKYC)"),
            ("Q12","CEO / CCO",
             "Do you have a mobile app or USSD service for customers? How many active users does it have, and what can they do on it (check balance, apply for loan, repay)?",
             "Active app with >20% of customer base using it = Solid. No app = Low.",
             "Low: No customer-facing app or USSD service\nPartial: App or USSD exists but limited functionality or low adoption\nSolid: App or USSD with loan application, repayment, and account view — actively used by a significant share of clients"),
        ]
    },
    {
        "title": "SECTION 4: Data, Analytics & Credit Scoring",
        "questions": [
            ("Q13","CEO / CTO",
             "Can you generate a portfolio quality report (e.g. PAR by branch, loan officer, product) on demand today — or does it require manual compilation?",
             "On-demand = Solid. Requires IT to pull manually = Low.",
             "Low: Reports require manual compilation, often by IT or finance team\nPartial: Standard reports available from CBS, but custom analysis requires manual work\nSolid: On-demand dashboards available to management and branch level, updated in real time"),
            ("Q14","CEO / CTO",
             "Do you use credit scoring in your lending decisions? If yes, is it rule-based, statistical, or machine learning-based? What data inputs are used?",
             "ML with alternative data = Solid. No scoring = Low. Scorecard rules = Partial.",
             "Low: No credit scoring; decisions based on loan officer judgment only\nPartial: Simple scorecard or rules-based scoring; internal repayment history only\nSolid: Statistical or ML-based scoring; uses alternative data (mobile money history, bureau data); >50% of decisions automated"),
            ("Q15","CEO / CTO",
             "Do you use data from external credit bureaus? What is the coverage of the bureau in your target market?",
             "Bureau integrated in decisioning = Solid. No bureau access = Low.",
             "Low: No credit bureau access or bureau coverage <20% of target clients\nPartial: Bureau data available but used informally or coverage is limited\nSolid: Credit bureau integrated into credit process, good coverage of target clients"),
            ("Q16","CEO / CTO",
             "How strongly are data privacy principles applied — do you have a data protection policy, and are you registered with the national data protection authority where required?",
             "Kenya DPA, Nigeria NDPA, Uganda DPA, SA POPIA, Philippines DPA, Indonesia PDPA etc. Non-compliance is a legal and reputational risk.",
             "Low: No data protection policy; not registered with authority\nPartial: Policy exists, registration in progress, partial compliance\nSolid: Registered with authority, documented policy, staff trained, active compliance"),
        ]
    },
    {
        "title": "SECTION 5: IT Governance & Team",
        "questions": [
            ("Q17","CEO / CTO",
             "Do you have a dedicated IT steering committee or digital transformation committee at management level? Does the board have visibility into technology decisions?",
             "Board-level tech committee = Solid. No governance structure = Low.",
             "Low: No IT governance structure; technology managed ad hoc\nPartial: IT committee at management level only\nSolid: IT steering committee at management level + board-level tech/digital strategy committee"),
            ("Q18","CEO / CTO",
             "How stable and skilled is the IT team? What is the average staff tenure in the IT department, and have you been able to hire people with modern skills (data, mobile, cloud) in recent years?",
             ">3 years tenure + modern skills = Solid. High turnover + no new skills = Low.",
             "Low: High turnover (<2 years average tenure), no recent hires with modern digital skills\nPartial: Moderate tenure (2–3 years), some skill gaps\nSolid: Stable team (>3 years), recent hires with relevant skills (cloud, data, mobile)"),
            ("Q19","CEO / CTO",
             "Have you worked with any external technology partners, technical assistance providers, or fintech companies to accelerate your digital agenda?",
             "TA from CGAP, IFC, Gates, donors, fintech partnerships = Solid",
             "Low: No external tech partnerships or TA engagement\nPartial: Some donor/TA engagement but not institutionalised\nSolid: Active fintech partnerships, TA programme, or formal technology partnership"),
        ]
    },
    {
        "title": "SECTION 6: Cybersecurity & Business Continuity",
        "questions": [
            ("Q20","CEO / CTO",
             "Has the institution experienced any significant cybersecurity incident in the past 3 years (data breach, fraud, ransomware)? If yes, how was it resolved?",
             "Clean record OR incident with documented remediation = Partial/Solid. Incident + no response = Low.",
             "Low: Significant incident with no documented response or remediation\nPartial: No major incident, but limited security controls in place\nSolid: No major incident AND strong controls in place (MFA, staff training, tested backups)"),
            ("Q21","CEO / CTO",
             "Do staff use multi-factor authentication (MFA) to access the core banking system and key internal systems?",
             "MFA is now a baseline security requirement.",
             "Low: No MFA on any system\nPartial: MFA on some systems\nSolid: MFA enforced across all critical systems including CBS"),
            ("Q22","CEO / CTO",
             "What is your backup and disaster recovery approach? How quickly can you restore systems if there is an outage?",
             "Cloud backup tested <4h RTO = Solid. No DR plan = Low.",
             "Low: No documented backup or disaster recovery plan\nPartial: Regular backups in place but recovery not tested\nSolid: Regular tested backups with documented recovery plan; target recovery <24 hours"),
        ]
    },
    {
        "title": "SECTION 7: External Environment (Context — not scored)",
        "context": True,
        "questions": [
            ("Q23","Desk Research",
             "What % of people in the country use a phone for financial transactions?",
             "Check GSMA/World Bank data. High mobile money penetration raises expectations.",
             "<30%: Limited digital enablement\n30–60%: Moderate\n>60%: High digital enablement"),
            ("Q24","CEO / COO",
             "What digital financial regulations most affect your institution (e.g. mobile transaction limits, eKYC rules, data hosting rules, agency banking regulations)?",
             "Understanding regulatory context matters for assessing ambition vs. reality.",
             "Restrictive regulatory environment\nSome enabling regulations\nStrong enabling framework"),
        ]
    },
]

# ─────────────────────────────────────────────────────────────
# DD QUESTION DATA
# ─────────────────────────────────────────────────────────────
DD_SECTIONS = [
    {
        "title": "DD SECTION 1: Digital Strategy & Governance",
        "questions": [
            ("DD1","CEO / COO",
             "Walk through your digital roadmap for the next 3 years. What are the 3-4 main initiatives, what is the total budget allocated, and what measurable outcomes are expected?",
             "Look for: budget specificity, KPIs, linkage to business goals. Vague answers = Low.",
             "Low: No clear roadmap or budget allocated\nPartial: Some initiatives identified, budget approximate\nSolid: Detailed roadmap with budget, timelines, and measurable KPIs"),
            ("DD2","CEO / COO",
             "How does the board approve and monitor technology investments — is there a separate technology budget line approved by the board?",
             "Shows whether tech is board-level priority or buried in ops",
             "Low: No separate tech budget; board has no visibility\nPartial: Tech budget exists but not specifically board-approved\nSolid: Board approves and monitors separate technology budget annually"),
            ("DD3","CEO",
             "Does the institution benchmark its digital maturity against peers? Have you participated in any formal digital assessment (e.g. CGAP Digital Readiness, IFC tools)?",
             "Shows self-awareness and improvement orientation",
             "Low: No benchmarking or external assessment\nPartial: Informal comparison or partial external assessment\nSolid: Regular benchmarking; has undergone formal external digital assessment"),
            ("DD4","CEO / CTO",
             "What is your philosophy on build vs. buy vs. partner for technology solutions? Give an example of each where possible.",
             "Shows sophistication of tech decision-making",
             "Low: All technology built in-house with no external partners\nPartial: Mix of approaches, but no clear framework\nSolid: Clear framework with examples; actively leverages SaaS/fintech partnerships"),
            ("DD5","CEO / CTO",
             "How do you measure the return on technology investments? Can you share an example of a tech investment, its cost, and the measurable outcome?",
             "Shows data-driven decision-making",
             "Low: No measurement of tech ROI\nPartial: Informal tracking, anecdotal outcomes\nSolid: Formal measurement; can cite specific example with cost and measurable outcome"),
            ("DD6","CEO",
             "Is digital transformation a standing agenda item at board meetings? Does the board include any member with a digital or technology background?",
             "Board competence on digital is increasingly important",
             "Low: Digital not on board agenda; no tech-savvy board members\nPartial: Occasional board discussion; limited tech expertise on board\nSolid: Regular board agenda item; at least one board member with tech/digital background"),
        ]
    },
    {
        "title": "DD SECTION 2: Core Banking System — Deep Dive",
        "questions": [
            ("DD7","CTO",
             "Name the CBS vendor, version, and deployment model (cloud SaaS / hosted / on-premise). Is it a subscription or perpetual license? Who provides support — vendor or in-house?",
             "Know the platform. Top platforms: Mambu, Temenos, Oradian, Musoni, Craft Silicon, Finacle, T24. In-house built = High Risk.",
             "Low: In-house built system, unknown vendor, or system no longer actively supported\nPartial: Known vendor, on-premise or hosted, standard platform\nSolid: Cloud-native SaaS from active vendor with ongoing development roadmap"),
            ("DD8","CTO",
             "What CBS modules are live today (loan origination, savings, teller, GL, HR, CRM)? Which are planned? Are all branches on the same system, or are some still on spreadsheets?",
             "Fragmented systems = Low. Unified platform = Solid.",
             "Low: Core modules only; some branches on spreadsheets\nPartial: Most modules on CBS but some gaps; all branches connected\nSolid: Full suite of modules live; all branches on unified real-time system"),
            ("DD9","CTO",
             "What is the CBS vendor's product roadmap? Is the product actively developed with regular updates, or is it legacy maintenance only? What is your contingency if the vendor ends support?",
             "Vendor viability is a key risk factor",
             "Low: Vendor has no active roadmap or unclear future\nPartial: Vendor active but roadmap limited\nSolid: Active vendor with clear product roadmap; institution has assessed migration contingency"),
            ("DD10","CTO",
             "Describe how your CBS connects to other systems — does it use real-time APIs or periodic file transfers? Walk through one live integration (e.g. mobile money reconciliation).",
             "Real-time API integrations = Solid. Manual file transfers or no integrations = Low.",
             "Low: All integrations via manual file transfers or none exist\nPartial: Some API integrations; mobile money partially integrated\nSolid: Multiple real-time API integrations; mobile money, bureau, and/or eKYC connected via API"),
            ("DD11","CTO",
             "What happens when the CBS goes down — is there a downtime playbook, and what is the average downtime per month?",
             "<1% downtime with playbook = Solid",
             "Low: No downtime playbook; outages handled reactively\nPartial: Some playbook; average downtime 1-4 hours/month\nSolid: Documented playbook; average downtime <1 hour/month; tested recovery procedures"),
            ("DD12","CTO",
             "Have you evaluated migrating to a cloud-native CBS? What are the main factors for or against in your context?",
             "This reveals strategic thinking and awareness of options",
             "Low: Has not considered cloud CBS\nPartial: Has evaluated but no clear plan\nSolid: Currently on cloud or has a concrete migration plan with timeline"),
        ]
    },
    {
        "title": "DD SECTION 3: Digital Channels & Mobile Money — Deep Dive",
        "questions": [
            ("DD13","CCO / CTO",
             "Provide the channel breakdown for loan applications received in the last 12 months: branch walk-in, field officer paper form, field officer digital app, customer self-service (app/USSD/web).",
             "Actual numbers required. Inability to answer = Partial/Low.",
             "Low: Unable to provide breakdown; all applications paper-based\nPartial: Some digital applications but majority still paper-based\nSolid: >30% applications through digital self-service channels; data readily available"),
            ("DD14","CCO / CTO",
             "List all mobile money integrations (provider name, transaction types supported, volume last 12 months, real-time API or batch).",
             "Depth and breadth of integration matters",
             "Low: No mobile money integration or integration only via manual reconciliation\nPartial: One integration, limited to one transaction type\nSolid: Multiple integrations (or main provider); both disbursement and repayment; real-time reconciliation"),
            ("DD15","CCO / CTO",
             "Do you operate an agent network? How many active agents, what can they do, and what technology do they use? How is agent fraud controlled?",
             "Agent networks extend reach but require strong controls",
             "Low: No agent network in market where it would be expected, or uncontrolled agents\nPartial: Small agent network, limited functionality, basic controls\nSolid: Active agent network with digital agent app, real-time monitoring, fraud controls"),
            ("DD16","CCO / CTO",
             "Do you have partnerships with any digital platforms or ecosystems — super-apps, e-commerce platforms, supply chain apps — for customer acquisition or embedded lending?",
             "Shows forward-looking distribution strategy",
             "Low: No platform partnerships\nPartial: Informal or nascent partnership\nSolid: Active partnership with measurable volume through digital platform"),
            ("DD17","CTO",
             "Describe your digital customer support — WhatsApp, in-app chat, call center, chatbot. What is the first-contact resolution rate?",
             "Digital-first support = Solid. Branch only = Low.",
             "Low: Support only available at branches\nPartial: Phone/email support available, no digital channels\nSolid: Multi-channel digital support (WhatsApp, app, chat) with tracking of resolution rates"),
        ]
    },
    {
        "title": "DD SECTION 4: Data, Analytics & AI/ML — Deep Dive",
        "questions": [
            ("DD18","CTO / CEO",
             "Describe your data architecture. Is your transactional data in the CBS only, or do you have a separate data warehouse or data lake? What tools do you use for analytics (Power BI, Tableau, Excel)?",
             "Separate data warehouse + BI tool = Solid. CBS only + Excel = Low.",
             "Low: No data warehouse; all analytics done in CBS or Excel\nPartial: Some consolidated reporting, basic BI tool\nSolid: Data warehouse or lake; dedicated BI platform; analytics available to management and branch level"),
            ("DD19","CTO / CEO",
             "Walk through your current credit scoring approach in detail. What data inputs, what model type, who built it, how often is it recalibrated, and what is the auto-approval rate?",
             "This is the most value-critical question for a lending institution.",
             "Low: No credit scoring; pure loan officer judgment\nPartial: Scorecard-based model, internal data only, infrequent updates\nSolid: Statistical or ML model; alternative data inputs; auto-approval >30%; regularly recalibrated with performance data"),
            ("DD20","CTO",
             "Have you deployed or piloted any AI or machine learning — for credit scoring, fraud detection, delinquency prediction, or customer segmentation? Describe the use case, model, and measurable outcome.",
             "Shows frontier digital maturity",
             "Low: No AI/ML whatsoever\nPartial: Pilot underway or evaluation stage\nSolid: Deployed ML in production; can demonstrate measurable improvement (lower defaults, higher approval rates, etc.)"),
            ("DD21","CTO",
             "Do you have mobile money transaction data from MNOs feeding into credit decisions? Do you have data-sharing agreements with MNOs or bureaus for alternative data?",
             "Alternative data = key differentiator",
             "Low: No alternative data sources\nPartial: Some alternative data used informally\nSolid: Formal data-sharing agreements; alternative data integrated into credit model"),
            ("DD22","CTO",
             "Do you have a formal data governance policy — data ownership, quality standards, access controls, retention/deletion schedule? Who is accountable for data quality?",
             "Data governance is fundamental infrastructure",
             "Low: No data governance policy\nPartial: Basic policy exists but not enforced\nSolid: Documented and enforced data governance; named data owner; regular data quality reviews"),
            ("DD23","CEO / CTO",
             "Are you registered with the national data protection authority? Do you have a Data Protection Officer? Have you assessed your AI/credit scoring tools for fairness and explainability (can you tell a customer why they were declined)?",
             "Regulatory requirement in Kenya, Nigeria, Uganda, SA, Philippines, Indonesia",
             "Low: Not registered, no DPO, no explainability mechanism\nPartial: Registration in progress; partial compliance\nSolid: Registered; DPO appointed; explainability built into credit model; staff trained on data privacy"),
        ]
    },
    {
        "title": "DD SECTION 5: API Architecture & Fintech Integration",
        "questions": [
            ("DD24","CTO",
             "Does your CBS expose APIs? What specification (REST/SOAP)? Is there API documentation? Can you describe an integration you have built on top of these APIs?",
             "REST API with documentation = Solid. No APIs = Low.",
             "Low: No API layer; all integrations manual\nPartial: Some APIs but limited documentation or spec\nSolid: Documented REST API; can demonstrate a live integration built on it"),
            ("DD25","CTO",
             "List your current third-party integrations and how they are achieved: credit bureaus, mobile money, eKYC providers, payment aggregators, insurance, fintech partners. Which are real-time API and which are file-transfer batch?",
             "Quality and depth of ecosystem integration matters",
             "Low: 0–1 integration, all file-based\nPartial: 2–3 integrations, mix of API and file\nSolid: 4+ real-time API integrations across multiple third-party systems"),
            ("DD26","CTO",
             "Are you integrated with your country's interoperable payment infrastructure (e.g. Ghana GHIPSS, Kenya Pesalink, Philippines InstaPay, Tanzania TIPS)?",
             "Shows ecosystem participation",
             "Low: Not connected to national payment infrastructure\nPartial: Connected to one element\nSolid: Actively integrated with national payment switch; real-time payments enabled"),
            ("DD27","CEO / CTO",
             "Do you have any Banking-as-a-Service or embedded finance arrangement — either providing services to other fintechs, or consuming fintech services to enhance your own products?",
             "Forward-looking revenue and distribution model",
             "Low: No BaaS or embedded finance consideration\nPartial: Exploring\nSolid: Active arrangement in place with measurable volume"),
        ]
    },
    {
        "title": "DD SECTION 6: Digital Lending & Process Automation",
        "questions": [
            ("DD28","CEO / CCO",
             "What is the average time from loan application to disbursement for your main product? What specific steps are still manual? What is the target?",
             "Time to cash is a key competitiveness metric",
             "Low: >2 weeks; most steps manual\nPartial: 3–14 days; partial automation\nSolid: <3 days; application to disbursement mostly automated; target is same-day for existing clients"),
            ("DD29","CEO / CTO",
             "Do you have any automated loan decisioning — where the system approves or declines without human review? For what products and loan sizes? How does PAR compare for auto-approved vs. manually approved loans?",
             "Auto-decisioning with good PAR = Solid",
             "Low: No automated decisioning whatsoever\nPartial: Some automation but below 20% of volume\nSolid: Auto-decisioning for >30% of loans; PAR comparable to manual portfolio"),
            ("DD30","CEO / CCO",
             "What is your early warning system for delinquency — are alerts automated or does detection rely on field officer visits?",
             "Automated EWS = Solid. Field visit only = Low.",
             "Low: Delinquency only detected at field visit\nPartial: Some automated alerts but not systematic\nSolid: Automated early warning system triggers alerts before repayment due date"),
            ("DD31","CEO / CCO",
             "Do you offer any fully digital products — where the full journey (application, approval, disbursement) requires no physical interaction? If yes, what is the loan size range and PAR?",
             "Fully digital product = proof of digital maturity",
             "Low: No fully digital product\nPartial: Digital product in pilot or for very small loans only\nSolid: Fully digital product in production with demonstrated performance data"),
        ]
    },
    {
        "title": "DD SECTION 7: Cybersecurity & Compliance",
        "questions": [
            ("DD32","CTO / CEO",
             "Do you have a documented Information Security Policy approved by the board? Which cybersecurity framework does it align to (NIST CSF, ISO 27001, CIS Controls, local regulation)?",
             "Framework alignment = Solid. No policy = Low.",
             "Low: No information security policy\nPartial: Policy exists but not board-approved or not aligned to a framework\nSolid: Board-approved policy; aligned to recognised framework (NIST CSF 2.0, ISO 27001, etc.)"),
            ("DD33","CTO",
             "Have you conducted external penetration testing or a security vulnerability assessment in the past 12 months? What was found, and how were critical findings addressed?",
             "Annual pentest = baseline. Findings remediated = Solid.",
             "Low: No external security testing\nPartial: Testing done >12 months ago or findings not remediated\nSolid: Annual external pentest; critical findings documented and remediated"),
            ("DD34","CTO",
             "Describe your patch management process — how quickly are critical security patches applied to the CBS, operating systems, and key applications?",
             "<30 days for critical patches = Solid",
             "Low: No patch management process\nPartial: Patches applied but inconsistently or slowly (>90 days)\nSolid: Defined patch management; critical patches applied within 30 days"),
            ("DD35","CTO",
             "Describe your backup and disaster recovery approach — frequency, where backups are stored (cloud/offsite), and when you last tested your recovery. What is your target recovery time?",
             "Tested recovery + offsite storage = Solid",
             "Low: No tested backup or DR plan\nPartial: Backups in place but not tested or stored on-site only\nSolid: Daily backups stored offsite/cloud; DR plan tested within last 12 months; RTO defined"),
            ("DD36","CEO / CTO",
             "Have you experienced digital fraud (mobile money fraud, SIM-swap, phishing of staff) in the past 3 years? What controls have been implemented to prevent recurrence?",
             "Shows security culture and responsiveness",
             "Low: Fraud incidents with no documented response or controls\nPartial: No recent incidents but limited preventive controls\nSolid: Proactive controls in place (staff training, transaction monitoring, MFA); incident response plan documented"),
        ]
    },
]


# ═══════════════════════════════════════════════════════════════
# SHEET 1 — INSTRUCTIONS
# ═══════════════════════════════════════════════════════════════
def build_instructions(wb):
    ws = wb.create_sheet("Instructions")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 60
    ws.column_dimensions["D"].width = 3

    r = 1

    # Main title
    c = merged_cell(ws, r, 1, r, 4)
    c.value = "ABLER NORDIC — DIGITAL MATURITY ASSESSMENT TOOL"
    c.fill = fill(AN_DARK_BLUE)
    c.font = afont(bold=True, color=WHITE, size=20)
    c.alignment = center_wrap()
    rh(ws, r, 44); r += 1

    c = merged_cell(ws, r, 1, r, 4)
    c.value = "For use during investment due diligence of Microfinance Institutions (MFIs)"
    c.fill = fill(AN_LIGHT)
    c.font = afont(italic=True, color=AN_DARK_BLUE, size=12)
    c.alignment = center_wrap()
    rh(ws, r, 24); r += 1
    rh(ws, r, 10); r += 1

    def sec_title(row, txt):
        c = merged_cell(ws, row, 1, row, 4)
        c.value = txt
        c.fill = fill(SEC_HEADER)
        c.font = afont(bold=True, color=WHITE, size=13)
        c.alignment = left_mid()
        rh(ws, row, 26)

    def body(row, label, text, bg=ROW_WHITE):
        cl = ws.cell(row=row, column=2)
        cl.value = label
        cl.font = afont(bold=True, color=AN_DARK_BLUE, size=11)
        cl.fill = fill(bg)
        cl.alignment = left_mid()
        ct = ws.cell(row=row, column=3)
        ct.value = text
        ct.font = afont(size=11)
        ct.fill = fill(bg)
        ct.alignment = left_top()
        rh(ws, row, 30)

    # Section 1
    sec_title(r, "1.  WHAT IS THIS TOOL?"); r += 1
    c = merged_cell(ws, r, 2, r, 3)
    c.value = (
        "This workbook is Abler Nordic's structured framework for evaluating the digital maturity "
        "of microfinance institutions (MFIs) during investment screening and due diligence. "
        "It ensures consistent, evidence-based assessment across all portfolio candidates and "
        "helps investment officers ask the right questions, interpret answers, and identify risks."
    )
    c.font = afont(size=11)
    c.fill = fill(ROW_WHITE)
    c.alignment = left_top()
    rh(ws, r, 55); r += 1
    rh(ws, r, 10); r += 1

    # Section 2
    sec_title(r, "2.  TWO-PHASE PROCESS"); r += 1
    phases = [
        ("PHASE 1 — Preliminary (Prelim)",
         "A 1-hour video call covering 22 scored questions across 6 sections, "
         "plus 2 context questions (not scored). Goal: quickly screen digital "
         "maturity and decide whether to proceed to full DD. "
         "Use the 'Prelim Assessment' sheet during the call."),
        ("PHASE 2 — Due Diligence (DD)",
         "A 1-hour deeper video call covering 36 scored questions across 7 "
         "sections. Conducted after reviewing the DD document pack (see "
         "'DD Documents Checklist'). Use the 'DD Assessment' sheet and review "
         "documents beforehand."),
    ]
    for i, (label, text) in enumerate(phases):
        body(r, label, text, bg=ROW_WHITE if i % 2 == 0 else ROW_ALT)
        r += 1
    rh(ws, r, 10); r += 1

    # Section 3
    sec_title(r, "3.  SCORING GUIDE"); r += 1
    scores = [
        ("0 — Low / None",
         "The institution has little or no capability in this area. "
         "Significant investment or remediation required.", LOW_BG),
        ("0.5 — Partial / Some",
         "Some foundations exist but gaps remain. Progress is being made "
         "but the institution is not yet fully capable.", ROW_ALT),
        ("1 — Solid / Full",
         "Strong capability clearly demonstrated. Evidence-based — "
         "ask for documents, data, or screen-shares to verify.", SOLID_BG),
    ]
    for label, text, bg in scores:
        cl = ws.cell(row=r, column=2)
        cl.value = label
        cl.font = afont(bold=True, size=11)
        cl.fill = fill(bg)
        cl.alignment = left_mid()
        ct = ws.cell(row=r, column=3)
        ct.value = text
        ct.font = afont(size=11)
        ct.fill = fill(bg)
        ct.alignment = left_top()
        rh(ws, r, 34); r += 1
    rh(ws, r, 10); r += 1

    # Section 4: colour legend
    sec_title(r, "4.  COLOUR LEGEND — SCORE INTERPRETATION"); r += 1
    legend = [
        (RED_SCORE, WHITE,
         "SCORE < 0.40 — RED (High Risk)",
         "High digital risk. Significant investment in technology may be needed "
         "before or alongside the investment."),
        (AMBER_SCORE, WHITE,
         "SCORE 0.40 – 0.69 — AMBER (Moderate)",
         "Digital foundations exist but gaps require attention. "
         "Consider TA or conditions attached to the investment."),
        (GREEN_BG, GREEN_TEXT,
         "SCORE >= 0.70 — GREEN (Strong)",
         "Strong digital maturity — above average for the peer group. "
         "Technology is an enabler rather than a risk."),
    ]
    for bg_col, fg_col, label, text in legend:
        cl = ws.cell(row=r, column=2)
        cl.value = label
        cl.font = afont(bold=True, color=fg_col, size=11)
        cl.fill = fill(bg_col)
        cl.alignment = center_wrap()
        ct = ws.cell(row=r, column=3)
        ct.value = text
        ct.font = afont(size=11)
        ct.fill = fill(ROW_WHITE)
        ct.alignment = left_top()
        rh(ws, r, 34); r += 1
    rh(ws, r, 10); r += 1

    # Section 5: who to involve
    sec_title(r, "5.  WHO TO INVOLVE IN EACH CALL"); r += 1
    who = [
        ("CEO / Managing Director",
         "Digital strategy, budget allocation, board dynamics, channel adoption, "
         "competitive positioning."),
        ("COO / Operations Director",
         "Process automation, field operations, agent networks, loan cycle, "
         "customer journeys."),
        ("CTO / IT Director",
         "Core banking system, API architecture, cybersecurity posture, "
         "data architecture, vendor management."),
        ("CCO / Chief Commercial Officer",
         "Digital channels, mobile money integration, customer acquisition "
         "models, digital product portfolio."),
    ]
    for i, (label, text) in enumerate(who):
        body(r, label, text, bg=ROW_WHITE if i % 2 == 0 else ROW_ALT)
        r += 1
    rh(ws, r, 10); r += 1

    # Section 6: tips
    sec_title(r, "6.  TIPS FOR VIDEO CALLS"); r += 1
    tips = [
        "Send the DD Documents Checklist (Sheet 4) at least 5 working days before the DD call.",
        "Ask for specific examples and data — do not accept general statements at face value.",
        "If the MFI cannot answer a quantitative question (e.g. % digital repayments), that is itself informative — score accordingly.",
        "Request screen shares of dashboards, CBS interfaces, or system reports where possible.",
        "Note any inconsistencies between what different participants say about the same system.",
        "Section 7 (External Environment) is not scored — use it to calibrate expectations for the country context.",
        "After each call, complete the Summary sheet while observations are fresh.",
        "Use the Notes column to record direct quotes or specific evidence supporting each score.",
    ]
    for i, tip in enumerate(tips):
        c = merged_cell(ws, r, 2, r, 3)
        c.value = f"  {i+1}.   {tip}"
        c.font = afont(size=11)
        c.fill = fill(ROW_WHITE if i % 2 == 0 else ROW_ALT)
        c.alignment = left_mid()
        rh(ws, r, 24); r += 1

    r += 1
    c = merged_cell(ws, r, 1, r, 4)
    c.value = "Abler Nordic AS  |  Digital Maturity Assessment Tool  |  Version 2026  |  Confidential"
    c.font = afont(italic=True, color="999999", size=10)
    c.alignment = center_wrap()
    rh(ws, r, 18)


# ═══════════════════════════════════════════════════════════════
# SHARED ASSESSMENT BUILDER
# ═══════════════════════════════════════════════════════════════
def build_assessment_sheet(wb, sheet_name, title, sections,
                            extra_input_row=False):
    ws = wb.create_sheet(sheet_name)
    ws.sheet_view.showGridLines = False
    set_col_widths(ws, [5, 18, 45, 45, 12, 15, 35])

    # Row 1: Title
    write_title_row(ws, 1, title, n_cols=7)

    # Row 2: Institution + Date
    c = merged_cell(ws, 2, 1, 2, 2)
    c.value = "Institution:"
    c.fill = fill(AN_LIGHT); c.font = afont(bold=True, color=AN_DARK_BLUE)
    c.alignment = right_mid()

    c = merged_cell(ws, 2, 3, 2, 4)
    c.fill = fill(ROW_WHITE)
    c.border = Border(bottom=Side(style="thin", color=AN_DARK_BLUE))

    ws.cell(row=2, column=5).value = "Date:"
    ws.cell(row=2, column=5).fill = fill(AN_LIGHT)
    ws.cell(row=2, column=5).font = afont(bold=True, color=AN_DARK_BLUE)
    ws.cell(row=2, column=5).alignment = right_mid()

    c = merged_cell(ws, 2, 6, 2, 7)
    c.fill = fill(ROW_WHITE)
    c.number_format = "DD-MMM-YYYY"
    rh(ws, 2, 22)

    if extra_input_row:
        c = merged_cell(ws, 3, 1, 3, 2)
        c.value = "Investment Officer:"
        c.fill = fill(AN_LIGHT); c.font = afont(bold=True, color=AN_DARK_BLUE)
        c.alignment = right_mid()
        c = merged_cell(ws, 3, 3, 3, 7)
        c.fill = fill(ROW_WHITE)
        c.border = Border(bottom=Side(style="thin", color=AN_DARK_BLUE))
        rh(ws, 3, 20)
        hdr_row = 4
    else:
        rh(ws, 3, 6)
        hdr_row = 4

    # Column headers
    write_col_headers(ws, hdr_row,
        ["#", "Who to Ask", "Question", "Scoring Guidance",
         "Answer\n(0 / 0.5 / 1)", "Score Label", "Notes"])

    current_row = hdr_row + 1
    answer_rows = []
    section_score_refs = []

    for sec in sections:
        is_ctx = sec.get("context", False)
        write_section_header(ws, current_row, sec["title"], context=is_ctx)
        current_row += 1

        q_refs = []
        for qi, (num, who, q, guidance, label) in enumerate(sec["questions"]):
            write_question_row(ws, current_row, num, who, q, guidance, label,
                               alt=(qi % 2 == 1), context=is_ctx)
            if not is_ctx:
                answer_rows.append(current_row)
                q_refs.append(f"E{current_row}")
                add_row_highlight_cf(ws, current_row)
            current_row += 1

        if not is_ctx:
            formula = f"=IFERROR(AVERAGE({','.join(q_refs)}),\"\")"
            ref = write_section_score_row(
                ws, current_row,
                f"  Section Score — {sec['title']}",
                formula)
            add_score_cf(ws, ref)
            section_score_refs.append(ref)
        else:
            # context note row
            c = merged_cell(ws, current_row, 1, current_row, 7)
            c.value = ("  Context only — use to calibrate interpretation "
                       "of scores above. Not included in overall score.")
            c.fill = fill(CONTEXT_BG)
            c.font = afont(italic=True, color=AN_DARK_BLUE, size=10)
            c.alignment = left_mid()
            rh(ws, current_row, 18)
            section_score_refs.append(None)

        current_row += 1
        rh(ws, current_row, 8)
        current_row += 1

    # Score summary
    write_score_summary(ws, current_row, sections, section_score_refs)

    # Dropdowns
    add_answer_dropdown(ws, answer_rows)

    # Freeze
    ws.freeze_panes = f"A{hdr_row + 1}"
    return ws


# ═══════════════════════════════════════════════════════════════
# SHEET 3 — PRELIM SUMMARY
# ═══════════════════════════════════════════════════════════════
def build_prelim_summary(wb):
    ws = wb.create_sheet("Prelim Summary")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 38
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 22
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 3

    r = 1
    c = merged_cell(ws, r, 1, r, 6)
    c.value = "PRELIMINARY ASSESSMENT — SUMMARY"
    c.fill = fill(AN_DARK_BLUE)
    c.font = afont(bold=True, color=WHITE, size=16)
    c.alignment = center_wrap()
    rh(ws, r, 36); r += 1

    rh(ws, r, 8); r += 1

    for label in ["Institution:", "Date:", "Assessed by:"]:
        cl = ws.cell(row=r, column=2)
        cl.value = label
        cl.font = afont(bold=True, color=AN_DARK_BLUE)
        cl.fill = fill(AN_LIGHT)
        cl.alignment = right_mid()
        cv = merged_cell(ws, r, 3, r, 5)
        cv.fill = fill(ROW_WHITE)
        cv.border = Border(bottom=Side(style="thin", color=AN_DARK_BLUE))
        rh(ws, r, 22); r += 1

    rh(ws, r, 10); r += 1

    # Section scores table
    c = merged_cell(ws, r, 2, r, 5)
    c.value = "SECTION SCORES"
    c.fill = fill(SEC_HEADER)
    c.font = afont(bold=True, color=WHITE, size=12)
    c.alignment = center_wrap()
    rh(ws, r, 22); r += 1

    # Table header
    hdr_data = [(2, "Section"), (3, "Score"), (4, "Rating"), (5, "")]
    ws.cell(row=r, column=2).value = "Section"
    ws.cell(row=r, column=2).font = afont(bold=True, color=AN_DARK_BLUE)
    ws.cell(row=r, column=2).fill = fill(AN_LIGHT)
    ws.cell(row=r, column=2).alignment = center_wrap()
    ws.cell(row=r, column=2).border = thin_border()
    ws.cell(row=r, column=3).value = "Score"
    ws.cell(row=r, column=3).font = afont(bold=True, color=AN_DARK_BLUE)
    ws.cell(row=r, column=3).fill = fill(AN_LIGHT)
    ws.cell(row=r, column=3).alignment = center_wrap()
    ws.cell(row=r, column=3).border = thin_border()
    c = merged_cell(ws, r, 4, r, 5)
    c.value = "Rating"
    c.font = afont(bold=True, color=AN_DARK_BLUE)
    c.fill = fill(AN_LIGHT)
    c.alignment = center_wrap()
    c.border = thin_border()
    rh(ws, r, 20); r += 1

    section_names = [
        ("1. Digital Strategy & Leadership", False),
        ("2. Core Banking System", False),
        ("3. Digital Channels & Mobile Money", False),
        ("4. Data, Analytics & Credit Scoring", False),
        ("5. IT Governance & Team", False),
        ("6. Cybersecurity & Business Continuity", False),
        ("7. External Environment (Context)", True),
    ]
    for i, (name, is_ctx) in enumerate(section_names):
        bg = CONTEXT_BG if is_ctx else (ROW_ALT if i % 2 else ROW_WHITE)
        cl = ws.cell(row=r, column=2)
        cl.value = name
        cl.fill = fill(bg)
        cl.font = afont(size=11, italic=is_ctx,
                        color=AN_DARK_BLUE if is_ctx else BLACK)
        cl.alignment = left_mid()
        cl.border = thin_border()

        sc = ws.cell(row=r, column=3)
        sc.value = "" if not is_ctx else "—"
        sc.fill = fill(bg)
        sc.font = afont(size=10, color="888888", italic=True)
        sc.alignment = center_wrap()
        sc.border = med_border()

        rc = merged_cell(ws, r, 4, r, 5)
        rc.value = "Context only" if is_ctx else ""
        rc.fill = fill(bg)
        rc.font = afont(size=10, italic=is_ctx)
        rc.alignment = center_wrap()
        rc.border = thin_border()
        rh(ws, r, 22); r += 1

    rh(ws, r, 10); r += 1

    # Key Observations
    c = merged_cell(ws, r, 2, r, 5)
    c.value = "KEY OBSERVATIONS FROM PRELIM CALL"
    c.fill = fill(SEC_HEADER)
    c.font = afont(bold=True, color=WHITE, size=12)
    c.alignment = center_wrap()
    rh(ws, r, 22); r += 1

    obs_start = r
    c = merged_cell(ws, r, 2, r+9, 5)
    c.value = ("Enter key observations, quotes from management, and "
               "supporting evidence here...")
    c.font = afont(size=11, color="888888", italic=True)
    c.fill = fill(ROW_WHITE)
    c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    c.border = Border(
        left=Side(style="medium", color=AN_DARK_BLUE),
        right=Side(style="medium", color=AN_DARK_BLUE),
        top=Side(style="medium", color=AN_DARK_BLUE),
        bottom=Side(style="medium", color=AN_DARK_BLUE))
    for i in range(10):
        rh(ws, r+i, 22)
    r += 10
    rh(ws, r, 10); r += 1

    # Red Flags
    c = merged_cell(ws, r, 2, r, 5)
    c.value = "RED FLAGS CHECKLIST"
    c.fill = fill(RED_SCORE)
    c.font = afont(bold=True, color=WHITE, size=12)
    c.alignment = center_wrap()
    rh(ws, r, 22); r += 1

    red_flags = [
        "No written digital strategy or board approval",
        "IT budget unknown or <2% of operating budget",
        "CBS is in-house built or on unidentified legacy system",
        "No mobile money integration in a market where MoMo is prevalent",
        "Unable to measure % of digital repayments",
        "No credit scoring — all decisions by loan officer judgment",
        "No credit bureau access",
        "No data protection policy / not registered with DPA",
        "No MFA on core banking system",
        "No disaster recovery plan or tested backup",
        "Significant cybersecurity incident with no documented remediation",
        "High IT staff turnover / no modern digital skills in team",
    ]
    for i, flag in enumerate(red_flags):
        bg = LOW_BG if i % 2 == 0 else "FFD0D0"
        cl = ws.cell(row=r, column=2)
        cl.value = "☐"
        cl.font = afont(size=13)
        cl.fill = fill(bg)
        cl.alignment = center_wrap()
        cl.border = thin_border()
        ct = merged_cell(ws, r, 3, r, 5)
        ct.value = flag
        ct.font = afont(size=10)
        ct.fill = fill(bg)
        ct.alignment = left_mid()
        ct.border = thin_border()
        rh(ws, r, 20); r += 1

    rh(ws, r, 10); r += 1

    # Recommendation
    c = merged_cell(ws, r, 2, r, 5)
    c.value = "RECOMMENDATION"
    c.fill = fill(AN_DARK_BLUE)
    c.font = afont(bold=True, color=WHITE, size=12)
    c.alignment = center_wrap()
    rh(ws, r, 22); r += 1

    recs = [
        ("☐  Proceed to Due Diligence",
         "Digital maturity is sufficient to warrant deeper assessment", SOLID_BG),
        ("☐  Request more information",
         "Specific gaps need clarification before proceeding", AN_LIGHT),
        ("☐  Decline / deprioritise",
         "Digital risk is too high for current investment thesis", LOW_BG),
    ]
    for rec, note, bg in recs:
        cl = ws.cell(row=r, column=2)
        cl.value = rec
        cl.font = afont(bold=True, size=11)
        cl.fill = fill(bg)
        cl.alignment = left_mid()
        cl.border = thin_border()
        ct = merged_cell(ws, r, 3, r, 5)
        ct.value = note
        ct.font = afont(size=10, italic=True)
        ct.fill = fill(bg)
        ct.alignment = left_mid()
        ct.border = thin_border()
        rh(ws, r, 26); r += 1


# ═══════════════════════════════════════════════════════════════
# SHEET 4 — DD DOCUMENTS CHECKLIST
# ═══════════════════════════════════════════════════════════════
def build_dd_checklist(wb):
    ws = wb.create_sheet("DD Documents Checklist")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 50
    ws.column_dimensions["C"].width = 38
    ws.column_dimensions["D"].width = 14

    r = 1
    c = merged_cell(ws, r, 1, r, 4)
    c.value = "DD DIGITAL MATURITY — DOCUMENTS REQUIRED"
    c.fill = fill(AN_DARK_BLUE)
    c.font = afont(bold=True, color=WHITE, size=16)
    c.alignment = center_wrap()
    rh(ws, r, 36); r += 1

    c = merged_cell(ws, r, 1, r, 4)
    c.value = ("Please send the following documents at least 5 working days before "
               "the scheduled Due Diligence meeting. If a document does not exist, "
               "the MFI should state this in writing.")
    c.fill = fill(AN_LIGHT)
    c.font = afont(italic=True, color=AN_DARK_BLUE, size=11)
    c.alignment = left_top()
    rh(ws, r, 36); r += 1
    rh(ws, r, 10); r += 1

    for ci, h in enumerate(["#", "Document", "Purpose", "Received?"], 1):
        c = ws.cell(row=r, column=ci)
        c.value = h
        c.fill = fill(AN_LIGHT)
        c.font = afont(bold=True, color=AN_DARK_BLUE, size=11)
        c.alignment = center_wrap()
        c.border = thin_border()
    rh(ws, r, 20); r += 1

    docs = [
        ("IT/Digital strategy document (latest version)",
         "Verify strategy quality and roadmap"),
        ("Core banking system documentation or product brochure",
         "Identify CBS vendor and capabilities"),
        ("IT organisational chart (with headcount by function)",
         "Assess team structure and capability"),
        ("IT budget (last 2 years, approved)",
         "Assess investment commitment"),
        ("List of all systems in use (CBS, MIS, HR, accounting, mobile apps) "
         "including: year implemented, vendor, cloud or on-premise, API availability",
         "Understand full tech stack"),
        ("System/network architecture diagram (if available)",
         "Understand integration architecture"),
        ("IT KPI report or dashboard (if available)",
         "Validate monitoring practices"),
        ("Data protection policy and registration certificate (if applicable)",
         "Confirm regulatory compliance"),
        ("Information security policy",
         "Assess cybersecurity governance"),
        ("Most recent management report with digital metrics "
         "(digital transactions %, app users, etc.)",
         "Validate digital channel data"),
        ("Credit scoring methodology document (if applicable)",
         "Understand credit model approach"),
        ("Most recent internal audit report (IT section)",
         "Identify known IT weaknesses"),
    ]
    for i, (doc, purpose) in enumerate(docs):
        bg = ROW_WHITE if i % 2 == 0 else ROW_ALT
        c = ws.cell(row=r, column=1)
        c.value = i + 1
        c.fill = fill(bg)
        c.font = afont(bold=True, color=AN_DARK_BLUE, size=11)
        c.alignment = center_wrap()
        c.border = thin_border()

        c = ws.cell(row=r, column=2)
        c.value = doc
        c.fill = fill(bg)
        c.font = afont(size=11)
        c.alignment = left_top()
        c.border = thin_border()

        c = ws.cell(row=r, column=3)
        c.value = purpose
        c.fill = fill(bg)
        c.font = afont(size=11, italic=True, color="444444")
        c.alignment = left_top()
        c.border = thin_border()

        c = ws.cell(row=r, column=4)
        c.value = "☐ Received"
        c.fill = fill(bg)
        c.font = afont(size=11)
        c.alignment = center_wrap()
        c.border = thin_border()
        rh(ws, r, 42); r += 1

    rh(ws, r, 10); r += 1
    c = merged_cell(ws, r, 1, r, 4)
    c.value = ("Note: Documents should be transmitted via secure channel. "
               "Receipt confirmation should be logged in the 'Received?' column.")
    c.fill = fill(AN_LIGHT)
    c.font = afont(italic=True, color=AN_DARK_BLUE, size=10)
    c.alignment = left_mid()
    rh(ws, r, 22)


# ═══════════════════════════════════════════════════════════════
# SHEET 6 — DD SUMMARY
# ═══════════════════════════════════════════════════════════════
def build_dd_summary(wb):
    ws = wb.create_sheet("DD Summary")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 3

    r = 1
    c = merged_cell(ws, r, 1, r, 6)
    c.value = "DUE DILIGENCE — DIGITAL MATURITY SUMMARY"
    c.fill = fill(AN_DARK_BLUE)
    c.font = afont(bold=True, color=WHITE, size=16)
    c.alignment = center_wrap()
    rh(ws, r, 36); r += 1
    rh(ws, r, 8); r += 1

    for label in ["Institution:", "Date:", "Investment Officer:",
                  "DD Call Participants:"]:
        cl = ws.cell(row=r, column=2)
        cl.value = label
        cl.font = afont(bold=True, color=AN_DARK_BLUE)
        cl.fill = fill(AN_LIGHT)
        cl.alignment = right_mid()
        cv = merged_cell(ws, r, 3, r, 5)
        cv.fill = fill(ROW_WHITE)
        cv.border = Border(bottom=Side(style="thin", color=AN_DARK_BLUE))
        rh(ws, r, 22); r += 1

    rh(ws, r, 10); r += 1

    # Combined scores
    c = merged_cell(ws, r, 2, r, 5)
    c.value = "COMBINED PRELIM + DD SECTION SCORES"
    c.fill = fill(SEC_HEADER)
    c.font = afont(bold=True, color=WHITE, size=12)
    c.alignment = center_wrap()
    rh(ws, r, 22); r += 1

    # Header
    ws.cell(row=r, column=2).value = "Section"
    ws.cell(row=r, column=2).font = afont(bold=True, color=AN_DARK_BLUE)
    ws.cell(row=r, column=2).fill = fill(AN_LIGHT)
    ws.cell(row=r, column=2).alignment = center_wrap()
    ws.cell(row=r, column=2).border = thin_border()
    for ci, h in [(3, "Prelim"), (4, "DD"), (5, "Combined")]:
        ws.cell(row=r, column=ci).value = h
        ws.cell(row=r, column=ci).font = afont(bold=True, color=AN_DARK_BLUE)
        ws.cell(row=r, column=ci).fill = fill(AN_LIGHT)
        ws.cell(row=r, column=ci).alignment = center_wrap()
        ws.cell(row=r, column=ci).border = thin_border()
    rh(ws, r, 20); r += 1

    combined_rows = [
        ("1. Digital Strategy & Leadership / Governance", True, True),
        ("2. Core Banking System", True, True),
        ("3. Digital Channels & Mobile Money", True, True),
        ("4. Data, Analytics & Credit Scoring / AI/ML", True, True),
        ("5. IT Governance & Team", True, True),
        ("6. Cybersecurity & Business Continuity / Compliance", True, True),
        ("7. API Architecture & Fintech Integration (DD only)", False, True),
        ("8. Digital Lending & Process Automation (DD only)", False, True),
        ("9. External Environment — Context (Prelim)", True, False),
    ]
    for i, (name, has_prelim, has_dd) in enumerate(combined_rows):
        bg = ROW_ALT if i % 2 else ROW_WHITE
        cl = ws.cell(row=r, column=2)
        cl.value = name
        cl.fill = fill(bg)
        cl.font = afont(size=11)
        cl.alignment = left_mid()
        cl.border = thin_border()
        for ci in [3, 4, 5]:
            sc = ws.cell(row=r, column=ci)
            if (ci == 3 and not has_prelim) or (ci == 4 and not has_dd):
                sc.value = "—"
            else:
                sc.value = ""
            sc.fill = fill(bg)
            sc.font = afont(size=10, color="888888", italic=True)
            sc.alignment = center_wrap()
            sc.border = thin_border()
        rh(ws, r, 22); r += 1

    rh(ws, r, 10); r += 1

    # Key Findings
    c = merged_cell(ws, r, 2, r, 5)
    c.value = "KEY FINDINGS & RED FLAGS"
    c.fill = fill(SEC_HEADER)
    c.font = afont(bold=True, color=WHITE, size=12)
    c.alignment = center_wrap()
    rh(ws, r, 22); r += 1

    c = merged_cell(ws, r, 2, r+11, 5)
    c.value = ("Document key findings, strengths, concerns, and red flags from "
               "both the Prelim and DD calls here. Reference specific question "
               "numbers where relevant.")
    c.font = afont(size=11, color="888888", italic=True)
    c.fill = fill(ROW_WHITE)
    c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    c.border = Border(
        left=Side(style="medium", color=AN_DARK_BLUE),
        right=Side(style="medium", color=AN_DARK_BLUE),
        top=Side(style="medium", color=AN_DARK_BLUE),
        bottom=Side(style="medium", color=AN_DARK_BLUE))
    for i in range(12):
        rh(ws, r+i, 22)
    r += 12
    rh(ws, r, 10); r += 1

    # Recommendation
    c = merged_cell(ws, r, 2, r, 5)
    c.value = "INVESTMENT RECOMMENDATION"
    c.fill = fill(AN_DARK_BLUE)
    c.font = afont(bold=True, color=WHITE, size=12)
    c.alignment = center_wrap()
    rh(ws, r, 22); r += 1

    recs = [
        ("☐  Proceed — digital maturity supports investment thesis",
         "Digital capability is sufficient and does not represent a material risk.",
         SOLID_BG),
        ("☐  Proceed with conditions — specific TA or remediation required",
         "Describe conditions below. Include in investment agreement and monitoring plan.",
         AN_LIGHT),
        ("☐  Further information required before decision",
         "Specify what additional information or clarification is needed.",
         "FFF2CC"),
        ("☐  Significant technology risk — consider carefully",
         "Fundamental technology weaknesses that may affect loan performance or scalability.",
         LOW_BG),
    ]
    for rec, note, bg in recs:
        cl = ws.cell(row=r, column=2)
        cl.value = rec
        cl.font = afont(bold=True, size=11)
        cl.fill = fill(bg)
        cl.alignment = left_mid()
        cl.border = thin_border()
        ct = merged_cell(ws, r, 3, r, 5)
        ct.value = note
        ct.font = afont(size=10, italic=True)
        ct.fill = fill(bg)
        ct.alignment = left_top()
        ct.border = thin_border()
        rh(ws, r, 30); r += 1

    rh(ws, r, 10); r += 1

    c = merged_cell(ws, r, 2, r, 5)
    c.value = "CONDITIONS / ADDITIONAL NOTES:"
    c.font = afont(bold=True, color=AN_DARK_BLUE, size=11)
    c.fill = fill(AN_LIGHT)
    c.alignment = left_mid()
    rh(ws, r, 20); r += 1

    c = merged_cell(ws, r, 2, r+7, 5)
    c.value = ""
    c.fill = fill(ROW_WHITE)
    c.border = Border(
        left=Side(style="medium", color=AN_DARK_BLUE),
        right=Side(style="medium", color=AN_DARK_BLUE),
        top=Side(style="medium", color=AN_DARK_BLUE),
        bottom=Side(style="medium", color=AN_DARK_BLUE))
    for i in range(8):
        rh(ws, r+i, 22)


# ═══════════════════════════════════════════════════════════════
# MAIN
# ═══════════════════════════════════════════════════════════════
def main():
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default sheet

    print("  Sheet 1: Instructions...")
    build_instructions(wb)

    print("  Sheet 2: Prelim Assessment...")
    build_assessment_sheet(
        wb,
        sheet_name="Prelim Assessment",
        title="PRELIMINARY DIGITAL MATURITY ASSESSMENT",
        sections=PRELIM_SECTIONS,
        extra_input_row=False
    )

    print("  Sheet 3: Prelim Summary...")
    build_prelim_summary(wb)

    print("  Sheet 4: DD Documents Checklist...")
    build_dd_checklist(wb)

    print("  Sheet 5: DD Assessment...")
    build_assessment_sheet(
        wb,
        sheet_name="DD Assessment",
        title="DUE DILIGENCE DIGITAL MATURITY ASSESSMENT",
        sections=DD_SECTIONS,
        extra_input_row=True
    )

    print("  Sheet 6: DD Summary...")
    build_dd_summary(wb)

    out = "/Users/steinar/Desktop/AN_Digital_Maturity_Assessment_2026.xlsx"
    print(f"\nSaving to: {out}")
    wb.save(out)
    print("File saved successfully.")

if __name__ == "__main__":
    main()
